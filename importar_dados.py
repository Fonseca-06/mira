import csv, json, urllib.request, urllib.error, datetime, sys, os
import openpyxl

SUPABASE_URL = 'https://kygpjnsqzhfcndqyuibw.supabase.co'
# Chave service_role lida do ambiente — NÃO versionar segredo.
# Rode com:  SUPABASE_SERVICE_KEY="sua_chave" python3 importar_dados.py
SERVICE_KEY = os.environ.get('SUPABASE_SERVICE_KEY')
if not SERVICE_KEY:
    sys.exit('Defina a variável de ambiente SUPABASE_SERVICE_KEY antes de rodar.')

BASE = '/mnt/c/Users/windows/Downloads'

def inserir(tabela, rows, batch=500):
    url = f'{SUPABASE_URL}/rest/v1/{tabela}'
    headers = {
        'apikey': SERVICE_KEY,
        'Authorization': f'Bearer {SERVICE_KEY}',
        'Content-Type': 'application/json',
        'Prefer': 'return=minimal'
    }
    total = 0
    erros = 0
    for i in range(0, len(rows), batch):
        lote = rows[i:i+batch]
        data = json.dumps(lote).encode('utf-8')
        req = urllib.request.Request(url, data=data, headers=headers, method='POST')
        try:
            with urllib.request.urlopen(req, timeout=60):
                total += len(lote)
                print(f'  lote {i//batch+1}: +{len(lote)} ({total}/{len(rows)})', flush=True)
        except urllib.error.HTTPError as e:
            msg = e.read().decode()
            print(f'  ERRO lote {i//batch+1}: {msg[:200]}')
            erros += 1
    return total, erros

def limpa_data(v):
    if not v or str(v).strip() == '' or str(v).strip() == '0.0':
        return None
    if isinstance(v, datetime.datetime):
        return v.date().isoformat()
    if isinstance(v, datetime.date):
        return v.isoformat()
    s = str(v).strip()
    # tenta converter dd/mm/yy ou dd/mm/yyyy
    for fmt in ('%d/%m/%Y', '%d/%m/%y', '%Y-%m-%d'):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None

def limpa_num(v):
    if v is None or str(v).strip() in ('', '0.0', '0'):
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None

def limpa_str(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

# ─── 1. PREÇOS HISTÓRICO (CSV) ────────────────────────────────────────────────

print('\n=== 1/4  Preços — histórico CSV ===')
precos = []
with open(f'{BASE}/04_precos_historico.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        preco = limpa_num(row.get('preco'))
        if not preco:
            continue
        precos.append({
            'origem':     limpa_str(row.get('origem')) or 'Meu',
            'medida':     limpa_str(row.get('medida')),
            'categoria':  limpa_str(row.get('categoria')),
            'segmento':   limpa_str(row.get('segmento')),
            'marca':      limpa_str(row.get('marca')),
            'descricao':  limpa_str(row.get('descricao')),
            'fornecedor': limpa_str(row.get('fornecedor')),
            'cliente':    limpa_str(row.get('cliente')),
            'uf':         limpa_str(row.get('uf')),
            'preco':      preco,
            'fonte':      limpa_str(row.get('fonte')),
            'data':       limpa_data(row.get('data')) or datetime.date.today().isoformat(),
            'obs':        limpa_str(row.get('obs')),
        })

print(f'  {len(precos)} registros carregados')
total, erros = inserir('precos', precos)
print(f'  ✓ {total} inseridos, {erros} erros')

# ─── 2. PREÇOS — PneuPrice_Base.xlsx (sheet Preços) ──────────────────────────

print('\n=== 2/4  Preços — PneuPrice_Base.xlsx ===')
wb = openpyxl.load_workbook(f'{BASE}/PneuPrice_Base.xlsx', read_only=True, data_only=True)
ws = wb['Preços']
rows_xl = list(ws.iter_rows(values_only=True))
wb.close()

precos2 = []
for row in rows_xl[1:]:
    if not row or not row[0]:
        continue
    preco = limpa_num(row[9])
    medida = limpa_str(row[1])
    if not preco or not medida:
        continue
    precos2.append({
        'origem':     limpa_str(row[0]) or 'Concorrente',
        'medida':     medida,
        'categoria':  limpa_str(row[2]),
        'segmento':   limpa_str(row[3]),
        'marca':      limpa_str(row[4]),
        'descricao':  limpa_str(row[5]),
        'fornecedor': limpa_str(row[6]),
        'cliente':    limpa_str(row[7]),
        'uf':         limpa_str(row[8]),
        'preco':      preco,
        'fonte':      limpa_str(row[10]),
        'data':       limpa_data(row[11]) or datetime.date.today().isoformat(),
        'obs':        None,
    })

print(f'  {len(precos2)} registros carregados')
total, erros = inserir('precos', precos2)
print(f'  ✓ {total} inseridos, {erros} erros')

# ─── 3. PREÇOS — Shopping_Precos_Pneus.xlsx (sheet Base de Dados) ─────────────

print('\n=== 3/4  Preços — Shopping_Precos_Pneus.xlsx ===')
wb = openpyxl.load_workbook(f'{BASE}/Shopping_Precos_Pneus.xlsx', read_only=True, data_only=True)
ws = wb['Base de Dados']
rows_xl = list(ws.iter_rows(values_only=True))
wb.close()

# colunas: Canal, Distribuidor/Fonte, UF, Data, Documento, Cliente, Municipio, Marca, Medida, Especificacao, Qtd, Preco unit.
precos3 = []
for row in rows_xl[1:]:
    if not row or not row[8]:
        continue
    preco = limpa_num(row[11])
    medida = limpa_str(row[8])
    if not preco or not medida:
        continue
    canal = limpa_str(row[0]) or ''
    obs_doc = limpa_str(row[4])
    precos3.append({
        'origem':     'Meu' if 'cantu' in canal.lower() or 'gripmaster' in canal.lower() else 'Concorrente',
        'medida':     medida,
        'categoria':  None,
        'segmento':   None,
        'marca':      limpa_str(row[7]),
        'descricao':  limpa_str(row[9]),
        'fornecedor': limpa_str(row[1]),
        'cliente':    limpa_str(row[5]),
        'uf':         limpa_str(row[2]),
        'preco':      preco,
        'fonte':      'NF Concorrente',
        'data':       limpa_data(row[3]) or datetime.date.today().isoformat(),
        'obs':        obs_doc,
    })

print(f'  {len(precos3)} registros carregados')
total, erros = inserir('precos', precos3)
print(f'  ✓ {total} inseridos, {erros} erros')

# ─── 4. CARTEIRA (CSV) ────────────────────────────────────────────────────────

print('\n=== 4/4  Carteira — 05_carteira.csv ===')
cadastros = []
with open(f'{BASE}/05_carteira.csv', encoding='utf-8-sig') as f:
    for row in csv.DictReader(f):
        razao = limpa_str(row.get('razao_social'))
        if not razao:
            continue
        cadastros.append({
            'cnpj':                limpa_str(row.get('cnpj')),
            'razao_social':        razao,
            'classificacao':       limpa_str(row.get('classificacao')),
            'cidade':              limpa_str(row.get('cidade')),
            'uf':                  limpa_str(row.get('uf')),
            'telefone':            limpa_str(row.get('telefone')),
            'email':               limpa_str(row.get('email')),
            'tipo_pessoa':         limpa_str(row.get('tipo_pessoa')),
            'cnae':                limpa_str(row.get('cnae')),
            'limite_credito':      limpa_num(row.get('limite_credito')),
            'data_cadastro':       limpa_data(row.get('data_cadastro')),
            'data_primeira_compra':limpa_data(row.get('data_primeira_compra')),
            'ultima_compra':       limpa_data(row.get('ultima_compra')),
            'status':              limpa_str(row.get('status')) or 'Ativo',
            'obs':                 limpa_str(row.get('obs')),
        })

print(f'  {len(cadastros)} registros carregados')
total, erros = inserir('cadastro', cadastros)
print(f'  ✓ {total} inseridos, {erros} erros')

# ─── 5. CATÁLOGO — extrair medidas e marcas únicas ───────────────────────────

print('\n=== Bonus: populando catálogo de medidas e marcas ===')

# Coleta todas as medidas e marcas dos dados importados
todas = precos + precos2 + precos3
med_map = {}  # medida → {categoria, aro}
marc_map = {}  # marca → segmento

for p in todas:
    m = p.get('medida')
    if m and m not in med_map:
        # Extrai aro da medida (ex: 185/60 R15 → 15)
        import re
        aro_match = re.search(r'[Rr](\d+)', m)
        aro = aro_match.group(1) if aro_match else None
        med_map[m] = {'medida': m, 'categoria': p.get('categoria'), 'aro': aro}

    ma = p.get('marca')
    if ma and ma not in marc_map:
        nossas = {'Speedmax','Itaro','Dynamo','Ni-Pon','Gripmaster'}
        origem = 'Nossa' if ma in nossas else ('Importado' if p.get('segmento','').startswith('IMP') else None)
        marc_map[ma] = {'marca': ma, 'segmento': p.get('segmento'), 'origem_marca': origem}

medidas_cat = list(med_map.values())
marcas_cat  = list(marc_map.values())

print(f'  {len(medidas_cat)} medidas únicas')
total, erros = inserir('ref_medidas', medidas_cat)
print(f'  ✓ {total} medidas inseridas')

print(f'  {len(marcas_cat)} marcas únicas')
total, erros = inserir('ref_marcas', marcas_cat)
print(f'  ✓ {total} marcas inseridas')

print('\n✅ Importação concluída!')
